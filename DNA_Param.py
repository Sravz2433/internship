import streamlit as st
import pandas as pd
import re
import csv
from io import StringIO, BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# --- Paste your process_single_dataset and format_excel_headers functions here ---
# (Copy the full code from your paste.txt for these two functions)
# For brevity, only the function calls are shown here; see your paste.txt for full code
def process_single_dataset(file_content):
    # Extract Sequence_ID
    sequence_id_match = re.match(r'>(\d+\.\d+)', file_content.strip())
    sequence_id = sequence_id_match.group(1) if sequence_id_match else "Unknown"

    # Extract intra and inter parts using regex
    intra_part = re.search(r'Intra-basepair parameters:(.*?)Inter-basepair parameters:', file_content, re.S)
    inter_part = re.search(r'Inter-basepair parameters:(.*)', file_content, re.S)

    if not intra_part or not inter_part:
        print("Invalid format: Missing intra or inter section.")
        return None

    # Intra parsing (robust)
    intra_lines = [line.strip() for line in intra_part.group(1).strip().split('\n') if line.strip()]
    intra_table_str = "\n".join(intra_lines)
    reader = csv.reader(StringIO(intra_table_str), delimiter='\t')
    intra_rows = list(reader)
    intra_header = [h.strip() for h in intra_rows[0]]
    intra_data = [[cell.strip() for cell in row] for row in intra_rows[1:] if any(cell.strip() for cell in row)]
    df_intra = pd.DataFrame(intra_data, columns=intra_header)

    for col in df_intra.columns:
        if col not in ["S.No", "Basepair"]:
            df_intra[col] = pd.to_numeric(df_intra[col], errors='coerce')

    # Build sequence string
    basepairs = df_intra["Basepair"].tolist()
    sequence = "".join(bp[0] for bp in basepairs if isinstance(bp, str) and bp)

    # Inter parsing (robust)
    inter_lines = [line.strip() for line in inter_part.group(1).strip().split('\n') if line.strip()]
    inter_table_str = "\n".join(inter_lines)
    reader = csv.reader(StringIO(inter_table_str), delimiter='\t')
    inter_rows = list(reader)
    inter_header = [h.strip() for h in inter_rows[0]]
    inter_data = [[cell.strip() for cell in row] for row in inter_rows[1:] if any(cell.strip() for cell in row)]
    df_inter = pd.DataFrame(inter_data, columns=inter_header)

    # Handle both "BP" and "BP step" as possible column names
    inter_exclude = set(["S.No", "BP", "BP step"])
    for col in df_inter.columns:
        if col not in inter_exclude:
            df_inter[col] = pd.to_numeric(df_inter[col], errors='coerce')

    # Build result row
    row = {
        "Sequence_ID": sequence_id,
        "Sequence": sequence
    }

    # Add intra parameters
    for param in [c for c in df_intra.columns if c not in ["S.No", "Basepair"]]:
        vals = df_intra[param].tolist()
        for i, v in enumerate(vals, 1):
            row[f"{param}_{i}"] = v
        row[f"{param}_AVG"] = pd.Series(vals, dtype='float').mean()

    # Add inter parameters
    for param in [c for c in df_inter.columns if c not in inter_exclude]:
        vals = df_inter[param].tolist()
        for i, v in enumerate(vals, 1):
            row[f"{param}_{i}"] = v
        row[f"{param}_AVG"] = pd.Series(vals, dtype='float').mean()

    return pd.DataFrame([row])



def format_excel_headers(filename, df):
    wb = load_workbook(filename)
    ws = wb.active
    ws.title = "Combined Data"
    if "Averages" in wb.sheetnames:
        del wb["Averages"]
    ws2 = wb.create_sheet("Averages")

    # Styles
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    inter_fill = PatternFill(start_color="CAE8BD", end_color="CAE8BD", fill_type="solid")
    intra_fill = PatternFill(start_color="F2E2B1", end_color="F2E2B1", fill_type="solid")
    blue_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    pink_fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
    green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    blue_font = Font(color="0070C0", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Format "Combined Data" sheet ---
    ws.insert_rows(1, 3)
    
    # First two columns: Sequence_ID and Sequence (merge rows 1 to 4)
    for col_num, (header, fill) in enumerate(
        [("Sequence_ID", pink_fill), ("Sequence", blue_fill)], start=1
    ):
        ws.merge_cells(start_row=1, start_column=col_num, end_row=4, end_column=col_num)
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = fill
        cell.font = bold_font
        cell.alignment = center_align

    # Get columns excluding the first two fixed columns
    col_names = list(df.columns)[2:]
    total_cols = len(col_names)

    # Group columns into intra and inter based on parameter names
    intra_params = {"Buckle", "Propeller", "Opening", "Shear", "Stretch", "Stagger"}
    inter_params = {"Tilt", "Roll", "Twist", "Shift", "Slide", "Rise"}
    groupings = []
    last_group = None
    group_start = 3
    for idx, col in enumerate(col_names, start=3):
        param_name = col.split('_')[0]
        group = "intra" if param_name in intra_params else "inter"
        if group != last_group:
            if last_group is not None:
                groupings.append((last_group, group_start, idx - 1))
            last_group = group
            group_start = idx
    groupings.append((last_group, group_start, 2 + total_cols))

    # Row 1: "Parameters" merged over all parameter columns
    if total_cols > 0:
        ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=2 + total_cols)
        param_cell = ws.cell(row=1, column=3, value="Parameters")
        param_cell.fill = intra_fill
        param_cell.font = bold_font
        param_cell.alignment = center_align

    # Row 2: Intra / Inter groups
    for group, start, end in groupings:
        label = "Intra Basepairs" if group == "intra" else "Inter Basepairs"
        if end > start:
            ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
        cell = ws.cell(row=2, column=start, value=label)
        cell.fill = inter_fill
        cell.font = bold_font
        cell.alignment = center_align

    # Row 3: Parameter groups (e.g., Buckle, Tilt, Roll)
    param_groupings = []
    last_param = None
    group_start = 3
    for idx, col in enumerate(col_names, start=3):
        param = col.split('_')[0]
        if param != last_param:
            if last_param is not None:
                param_groupings.append((last_param, group_start, idx - 1))
            last_param = param
            group_start = idx
    param_groupings.append((last_param, group_start, 2 + total_cols))

    for param, start, end in param_groupings:
        if end > start:
            ws.merge_cells(start_row=3, start_column=start, end_row=3, end_column=end)
        cell = ws.cell(row=3, column=start, value=param)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_align

    # Row 4: Subheaders (numbers and AVG)
    for idx, col in enumerate(col_names, start=3):
        param_name = col.split('_')[0]
        subheader = col.split('_')[1]
        fill = green_fill if param_name in intra_params else blue_fill
        cell = ws.cell(row=4, column=idx, value=subheader)
        cell.fill = fill
        cell.font = bold_font
        cell.alignment = center_align
        if subheader == "AVG":
            for row in range(5, 5 + df.shape[0]):
                data_cell = ws.cell(row=row, column=idx)
                data_cell.font = blue_font

    # Adjust column widths for "Combined Data"
    for idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=1):
        max_length = 0
        for cell in col_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(idx)].width = max_length + 2

    # --- Format "Averages" sheet ---
    # Find columns ending with '_AVG'
    avg_cols = [col for col in df.columns if col.endswith('_AVG')]

    # Write headings in row 1
    ws2.cell(row=1, column=1, value="Sequence ID").font = bold_font
    ws2.cell(row=1, column=1).alignment = center_align
    ws2.cell(row=1, column=1).border = thin_border

    ws2.cell(row=1, column=2, value="Sequence").font = bold_font
    ws2.cell(row=1, column=2).alignment = center_align  
    ws2.cell(row=1, column=2).border = thin_border

    for col_idx, col_name in enumerate(avg_cols, start=3):
        cell = ws2.cell(row=1, column=col_idx, value=col_name)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = header_fill

    # Write sequence IDs, sequences, and AVG values for each row
    for row_idx in range(df.shape[0]):
        # Sequence ID
        cell_id = ws2.cell(row=row_idx+2, column=1, value=df.iloc[row_idx, 0])
        cell_id.fill = pink_fill
        cell_id.border = thin_border
        # Sequence
        cell_seq = ws2.cell(row=row_idx+2, column=2, value=df.iloc[row_idx, 1])
        cell_seq.fill = blue_fill
        cell_seq.border = thin_border
        # AVG columns
        for col_idx, col_name in enumerate(avg_cols, start=3):
            cell = ws2.cell(row=row_idx+2, column=col_idx, value=df.iloc[row_idx][col_name])
            cell.border = thin_border

    # Ensure all cells in the used range have borders (including empty cells)
    total_cols = 2 + len(avg_cols)
    for row in range(1, df.shape[0] + 2):  # +2 for header row
        for col in range(1, total_cols + 1):
            ws2.cell(row=row, column=col).border = thin_border

    # Adjust column widths for better readability
    for col_cells in ws2.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        ws2.column_dimensions[col_cells[0].column_letter].width = max_length + 2

    wb.save(filename)

# ... (Insert your process_single_dataset and format_excel_headers functions here) ...

def process_uploaded_file(uploaded_file):
    file_content = uploaded_file.read().decode("utf-8")
    # Split the file into individual datasets, each starting with '>'
    datasets = re.split(r'(?=>)', file_content)
    all_results = []
    for ds in datasets:
        ds = ds.strip()
        if not ds:
            continue
        result = process_single_dataset(ds)
        if result is not None:
            all_results.append(result)
    if all_results:
        results_df = pd.concat(all_results, ignore_index=True)
        return results_df
    else:
        return None

def to_excel_download(df):
    output = BytesIO()
    temp_filename = "temp_results.xlsx"
    df.to_excel(output, index=False, engine="openpyxl")
    output.seek(0)
    # Now format headers in the Excel file
    with open(temp_filename, "wb") as f:
        f.write(output.read())
    format_excel_headers(temp_filename, df)
    with open(temp_filename, "rb") as f:
        formatted = f.read()
    return formatted

# --- Streamlit UI ---
st.title("DNA Parameter Parser & Excel Formatter")

st.markdown("""
Upload your DNA parameter `.txt` file (e.g., `out_Chr17_5nt_A.txt`).  
The app will parse, process, and let you download a formatted Excel file.
""")

uploaded_file = st.file_uploader("Choose a .txt file", type="txt")

if uploaded_file is not None:
    with st.spinner("Processing file..."):
        df = process_uploaded_file(uploaded_file)
    if df is not None:
        st.success(f"Processed {df.shape[0]} dataset(s).")
        st.dataframe(df.head(10))  # Show a preview
        excel_bytes = to_excel_download(df)
        st.download_button(
            label="Download Formatted Excel",
            data=excel_bytes,
            file_name="dna_parameters.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.error("No valid data was processed. Please check your file format.")

st.markdown("---")
st.markdown("Developed with ❤️ using Streamlit.")
